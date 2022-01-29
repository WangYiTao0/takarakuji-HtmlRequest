from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill

import matplotlib.pyplot as plt
import readcsv
from GenerateCSVFromUrl import GenerateCSV, remove_specific_row_from_csv, csv_to_xlsx
from readcsv import ReadCSVFile
import openpyxl

import matplotlib

matplotlib.rcParams['font.sans-serif'] = ['SimHei']  # 用黑体显示中文
matplotlib.rcParams['axes.unicode_minus'] = False
import numpy as np
import pandas as pd

url = 'http://www.ohtashp.com/topics/takarakuji/'
csv_file_Name = 'kuji.csv'
excel_file_Name = 'kuji.xlsx'
excel_file_temp_Name = 'kuji_temp.xlsx'

GenerateCSV(url, csv_file_Name)
ReadCSVFile(csv_file_Name)


# csv_to_xlsx(csv_file_Name,excel_file_Name,"2019-2021")

# for i in range(len(readcsv.key)):
#     print(readcsv.key[i])
#     print(kujiData[readcsv.key[i]])

# workbook
# worksheet
# row column cell
wb = openpyxl.load_workbook(excel_file_Name)

column = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']

row1 = ['回別', '抽選日', '本数字', 'bonus数字']

markFont = Font(name='Calibri', size=11, bold=True)


redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')
blueFill = PatternFill(start_color='0000FFFF',
                       end_color='000000FF',
                       fill_type='solid')
greenFill = PatternFill(start_color='0000FF00',
                       end_color='0000FF00',
                       fill_type='solid')
yellowFill = PatternFill(start_color='00FFFF00',
                       end_color='00FFFF00',
                       fill_type='solid')
purpleFill = PatternFill(start_color='00FF00FF',
                       end_color='00FF00FF',
                       fill_type='solid')
darkblueFill = PatternFill(start_color='000000FF',
                       end_color='000000FF',
                       fill_type='solid')
brown = PatternFill(start_color='00800000',
                       end_color='00800000',
                       fill_type='solid')

Fills = [redFill,blueFill,greenFill,yellowFill,purpleFill,darkblueFill,brown]

def drawAtariBg(ws):
    ws_AtariNums = ws
    # Change the name of sheet

    for i in range(1, len(readcsv.key) + 1):
        # 1 到 9  j - 1 0 到 8
        for j in range(0, 9):
            ws_AtariNums.cell(column=1, row=i).value = readcsv.key[i - 1]
            ws_AtariNums.cell(column=1, row=i).font = markFont
            ws_AtariNums.cell(column=2, row=i).value = readcsv.date[i - 1]
            ws_AtariNums.cell(column=j + 3, row=i).value = int(readcsv.num_str_dict[readcsv.key[i - 1]][j])
    ws_AtariNums.insert_rows(1)
    # ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)

    for cell in list(ws_AtariNums.rows)[0]:
        cell.fill = redFill
        cell.font = markFont
    ws_AtariNums['A1'] = row1[0]
    ws_AtariNums['B1'] = row1[1]
    ws_AtariNums['C1'] = row1[2]
    ws_AtariNums[get_column_letter(10) + '1'] = row1[3]
    ws_AtariNums.merge_cells(start_row=1, end_row=1, start_column=3, end_column=9)
    ws_AtariNums.merge_cells(start_row=1, end_row=1, start_column=10, end_column=11)


def DrawAtariNums():
    ws_atari = wb.active
    ws_atari.title = '当選数字'
    drawAtariBg(ws_atari)


DrawAtariNums()


def DrawBG(ws):
    keyLen = len(readcsv.key)
    for i in range(1, 38):
        # 1 - len(key)  1到153 -1 0 -152
        for j in range(0, keyLen):
            ws.cell(column=1, row=j + 1).value = readcsv.key[j]
            # 填满1到37 的数字
            ws.cell(column=i + 1, row=j + 1).value = i
def drawZouShiTu():
    ws_zhoushitu = wb.create_sheet("走势图")
    # 1- 37 column

    DrawBG(ws_zhoushitu)

    for i in range(0, len(readcsv.key)):
        kujiNums = readcsv.num_int_dict[readcsv.key[i]]
        for j in range(0, 7):
            ws_zhoushitu.cell(column=kujiNums[j] + 1, row=i + 1).fill = Fills[j]
        #for j in range(7, 9):
            # ws_zhoushitu.cell(column=kujiNums[j] + 1, row=i + 1).fill = blueFill

    wb.active = ws_zhoushitu

drawZouShiTu()


# 查看某一期 和 其前y期的数字重合
def CheckRepeatNum(x, y, sheet_name):
    ws_checkReqeatNum = wb.create_sheet(sheet_name)
    drawAtariBg(ws_checkReqeatNum)
    # DrawBG(ws_checkReqeatNum)

    # 要查的这期
    repeatNum = []
    totalReqeatCount = 0
    repeatNumDic = {}

    nums = readcsv.num_int_dict[readcsv.key[x]]
    # print(nums)
    for c in range(0, 7):
        ws_checkReqeatNum.cell(column=c + 3, row=x + 2).fill = redFill
    # 历遍后面 y 期
    for xIndex in range(0, 7):
        repeatCount = 0
        repeatNumDic[nums[xIndex]] = repeatCount
        for i in range(x + 1, x + y + 1):
            checkNums = readcsv.num_int_dict[readcsv.key[i]]
            # 双循环 查重复
            for yIndex in range(0, 7):
                if nums[xIndex] == checkNums[yIndex]:
                    repeatCount += 1
                    totalReqeatCount += 1
                    repeatNumDic[nums[xIndex]] = repeatCount
                    ws_checkReqeatNum.cell(column=yIndex + 3, row=i + 2).fill = blueFill

    #print(f'最近第{x}期 和 其前{y}期')
    #print(repeatNumDic)
    #print(totalReqeatCount)
    return totalReqeatCount;
    ws_checkReqeatNum.sheet_view.zoomScale = 90
    wb.active = ws_checkReqeatNum

def CheckReqeatNumCountFrequencey(x,excelName) :
    '''
    检测重复数字个数 每x期
    :param x:
    :param celName:
    :return:
    '''
    totalRepeatCount_list = []
    for i in range(50):
        totalRepeatCount_list.append(CheckRepeatNum(i, x, f"第{x}期和其前1期"))

    # converting list to array
    np_totalRepeatCount_list_array = np.array(totalRepeatCount_list)
    np_totalRepeatCount_list_array.sort()

    m = np.r_[True, np_totalRepeatCount_list_array[:-1] != np_totalRepeatCount_list_array[1:], True]
    counts = np.diff(np.flatnonzero(m))
    unq = np_totalRepeatCount_list_array[m[:-1]]
    print(np.c_[unq,counts])

    df = pd.DataFrame(np.c_[unq,counts])
    filepath = f'{excelName}.xlsx'

    df.to_excel(filepath, index=False)




#CheckReqeatNumCountFrequencey(1,'dif1')
#CheckReqeatNumCountFrequencey(2,'dif2')
#CheckReqeatNumCountFrequencey(3,'dif3')

# 奇 偶
def JiOu(num) :
    ji = np.zeros(num)
    ou = np.zeros(num)


    for i in range(num):
        nums = readcsv.num_int_dict[readcsv.key[i]]
        newNums = np.array(nums)
        newNums = np.delete(newNums,[7,8])
        for n in newNums:
            if n % 2 == 0:
                ou[i] += 1
            elif n % 2 == 1:
                ji[i] += 1
    plt.title("奇数走势")
    plt.plot(ji)
    plt.show()
    plt.title("偶数走势")
    plt.plot(ou)
    plt.show()

    for i in range(num):
        print("奇数 : " + str( ji[i]) + " + " + "偶数 : " + str(ou[i]))

def DaXiaoQu(num) :
    da = np.zeros(num)
    xiao = np.zeros(num)
    for i in range(num):
        nums = readcsv.num_int_dict[readcsv.key[i]]
        newNums = np.array(nums)
        newNums = np.delete(newNums,[7,8])
        for n in newNums:
            if n >= 19:
                da[i] += 1
            elif n <= 18:
                xiao[i] += 1

    plt.title("小区走势")
    plt.plot(xiao)
    plt.show()
    plt.title("大区走势")
    plt.plot(da)
    plt.show()

    for i in range(num):
        print("小区 : " + str(xiao[i]) + " + " + "大区 : " + str(da[i]))

def SanQu(num) :
    pass

JiOu(50)
DaXiaoQu(50)

'''
[[ 0 14]
 [ 1 18]
 [ 2 16]
 [ 3  2]]
[[ 0  1]
 [ 1 12]
 [ 2 15]
 [ 3 10]
 [ 4 10]
 [ 6  1]
 [ 7  1]]
[[ 0  1]
 [ 1  1]
 [ 2 11]
 [ 3  7]
 [ 4 12]
 [ 5 10]
 [ 6  6]
 [ 7  1]
 [ 8  1]]
 
 

 02	07	12	15	21	23	30	
 02	21	26	27	32	33	34	
 07	21	24	29	32	36	37
'''

# x = np.arange(0, 151, 1)  # 横坐标数据为从0到151之间，步长为1的等差数组
# y = np_totalRepeatCount_list_array[x]
#
# # 生成图形
# plt.plot(x, y)
# # 显示图形
# plt.show()

wb.save(excel_file_temp_Name)
